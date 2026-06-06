"""
avaliar_pipeline.py — Harness de avaliação multi-modelo (ROUBO)
================================================================
Lê o gabarito (planilha auditada) e a saída de UMA configuração da pipeline
(CSV gerado por pipeline_anonimizacao_roubo.py) e calcula, a NÍVEL DE ENTIDADE
POR TIPO, contra as 8 colunas do gabarito:

  - TP  = entidade do gabarito que a pipeline removeu (no tipo certo)
  - FN  = estava no gabarito e a pipeline deixou passar (vazamento — erro grave)
  - FP  = a pipeline removeu algo que não estava naquele tipo (sobre-remoção)
  - Accuracy = TP/(TP+FP+FN) ; Precision ; Recall ; F1 — por tipo, micro e macro

DUPLA VISÃO (decisão do usuário):
  (a) F1 por tipo — utilidade/atribuição de tipo.
  (b) Recall de SEGURANÇA — "removido sob QUALQUER placeholder = não vazou";
      mede o risco real de vazamento de PII (a métrica de segurança em destaque).

DE-PARA + HIGIENE (decisão do usuário): placeholders institucionais não entram
nas métricas de PII pessoal; e os CÓDIGOS INSTITUCIONAIS que o gabarito agrupou
na coluna Documento (IML, NOC, BOPM, processo/CNJ, boletim) são RECLASSIFICADOS
como higiene e retirados do gold de Documento (contados à parte).

Casamento robusto: minúsculas, sem acento, sem pontuação/espaço, com tolerância
a contenção (o gabarito é verbatim do texto descaracterizado).

Saídas:
  - relatório no console
  - resultados_avaliacao/<tag>_por_tipo.csv      (métricas por tipo)
  - resultados_avaliacao/<tag>_matriz_erros.csv  (gold x placeholder de remoção)
  - resultados_avaliacao/<tag>_top_fn.csv        (BOs com mais vazamento)
  - resultados_avaliacao/tabela_comparativa.csv  (1 linha por configuração; append)

Uso:
  python avaliar_pipeline.py --saida saidas/ner_spacy.csv
  python avaliar_pipeline.py --saida saidas/llm_qwen3-8b.csv
"""
import argparse
import csv
import json
import os
import re

import yaml
import openpyxl

from estagio1_regex import normalizar

# 8 colunas do gabarito (PII de pessoa física)
COLUNAS_PII = [
    'Nome(s)', 'Localizações (s)', 'Documento(s)', 'Emplacamento (s)',
    'Telefone (s)', 'Email (s) e contas (ex: instagram)', 'N° da VTR',
    'Vulgo(s)/Apelido(s)',
]

# categoria do Estágio 2 -> coluna do gabarito
CAT_ESTAGIO2 = {
    'nomes': 'Nome(s)',
    'locais': 'Localizações (s)',
    'estabelecimentos': 'Localizações (s)',
    'vulgos': 'Vulgo(s)/Apelido(s)',
}

# Reclassificação de Documento: gold que é CÓDIGO INSTITUCIONAL (não PII pessoal)
RE_DOC_INSTITUCIONAL = re.compile(
    r'(?i)\b(?:IML|NOC|BOPM|BO|RE|processo|CNJ|boletim|of[íi]cio|inqu[ée]rito|protocolo)\b'
)


def split_gold(valor):
    if valor is None:
        return []
    out = []
    for parte in str(valor).split(';'):
        n = normalizar(parte)
        if n:
            out.append((parte.strip(), n))
    return out


def casa(gn, preds):
    for p in preds:
        if not p:
            continue
        if gn == p:
            return True
        if len(gn) >= 4 and len(p) >= 4 and (gn in p or p in gn):
            return True
    return False


def carregar_gabarito(cfg):
    wb = openpyxl.load_workbook(cfg['paths']['dataset'], read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header) if h}
    gab = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq = row[idx[cfg['paths']['coluna_seq']]]
        gab[str(seq)] = {col: split_gold(row[idx[col]]) for col in COLUNAS_PII}
    return gab


def construir_de_para(cfg):
    """placeholder (sem colchetes) -> coluna do gabarito; e set de higiene."""
    ph2col = {}
    for ph, col in cfg['de_para'].items():
        ph2col[ph.strip('[]')] = col
    higiene = {h.strip('[]') for h in cfg['placeholders_higiene']}
    return ph2col, higiene


def predicoes_do_registro(reg, ph2col):
    """Constrói pred_por_coluna (normalizado) e todos_os_preds (visão segurança)."""
    e1 = json.loads(reg['removidos_estagio1_json'] or '{}')
    e2 = json.loads(reg['removidos_estagio2_json'] or '{}')
    pred = {c: [] for c in COLUNAS_PII}
    todos = []
    # Estágio 1
    for ph, spans in e1.items():
        norm = [normalizar(s) for s in spans]
        todos += norm
        col = ph2col.get(ph)
        if col in pred:
            pred[col] += norm
    # Estágio 2
    for cat, ents in e2.items():
        norm = [normalizar(s) for s in ents]
        todos += norm
        col = CAT_ESTAGIO2.get(cat)
        if col in pred:
            pred[col] += norm
    return pred, todos


def reclassificar_documento(gold_doc):
    """Separa Documento em (pii_pessoal, institucional)."""
    pii, inst = [], []
    for orig, gn in gold_doc:
        if RE_DOC_INSTITUCIONAL.search(orig):
            inst.append((orig, gn))
        else:
            pii.append((orig, gn))
    return pii, inst


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--saida', required=True, help='CSV gerado pela pipeline')
    ap.add_argument('--config', default='config.yaml')
    args = ap.parse_args()
    cfg = yaml.safe_load(open(args.config, encoding='utf-8'))
    res_dir = cfg['paths']['resultados_dir']
    os.makedirs(res_dir, exist_ok=True)
    tag = os.path.splitext(os.path.basename(args.saida))[0]

    gab = carregar_gabarito(cfg)
    ph2col, higiene = construir_de_para(cfg)

    # acumuladores
    tp = {c: 0 for c in COLUNAS_PII}
    fn = {c: 0 for c in COLUNAS_PII}
    fp = {c: 0 for c in COLUNAS_PII}
    gold_tot = {c: 0 for c in COLUNAS_PII}
    leak = {c: 0 for c in COLUNAS_PII}            # visão de segurança
    inst_total = inst_removidos = 0               # códigos institucionais (à parte)
    # matriz de erros: para FN, sob qual placeholder (se algum) foi removido
    matriz = {c: {'vazou': 0, 'mistype': 0} for c in COLUNAS_PII}
    top_fn = []                                   # (n_fn, seq)
    latencias, ptoks, ctoks = [], [], []
    n = 0

    with open(args.saida, encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for reg in reader:
            seq = str(reg['Seq'])
            if seq not in gab:
                continue
            n += 1
            latencias.append(float(reg.get('latencia_s') or 0))
            ptoks.append(int(reg.get('prompt_tokens') or 0))
            ctoks.append(int(reg.get('completion_tokens') or 0))

            pred, todos = predicoes_do_registro(reg, ph2col)
            fn_neste = 0
            for col in COLUNAS_PII:
                gold = gab[seq][col]
                if col == 'Documento(s)':
                    gold, inst = reclassificar_documento(gold)
                    inst_total += len(inst)
                    for orig, gn in inst:
                        if casa(gn, todos):
                            inst_removidos += 1
                gold_tot[col] += len(gold)
                preds = pred[col]
                casados = set()
                for orig, gn in gold:
                    if casa(gn, preds):
                        tp[col] += 1
                        casados.update(p for p in preds if gn == p or
                                       (len(gn) >= 4 and len(p) >= 4 and (gn in p or p in gn)))
                    else:
                        fn[col] += 1
                        fn_neste += 1
                        # vazou de fato ou só foi removido sob outro placeholder?
                        if casa(gn, todos):
                            matriz[col]['mistype'] += 1
                        else:
                            matriz[col]['vazou'] += 1
                            leak[col] += 1
                    # segurança: independ. do tipo
                    if not casa(gn, todos):
                        pass  # já contabilizado em leak quando FN; TP de segurança implícito
                for p in set(preds):
                    if p and p not in casados:
                        fp[col] += 1
            top_fn.append((fn_neste, seq))

    # -------------------- métricas --------------------
    def met(c):
        t, f, p = tp[c], fn[c], fp[c]
        rec = t / (t + f) if t + f else 0.0
        prec = t / (t + p) if t + p else 0.0
        f1 = 2 * rec * prec / (rec + prec) if rec + prec else 0.0
        acc = t / (t + f + p) if (t + f + p) else 0.0
        return rec, prec, f1, acc

    print("=" * 86)
    print(f"AVALIAÇÃO | config = {tag} | registros = {n}")
    print("=" * 86)
    print(f"{'Tipo':<36}{'Gold':>5}{'TP':>5}{'FN':>5}{'FP':>6}"
          f"{'Rec':>7}{'Prec':>7}{'F1':>7}{'Acc':>7}{'SegRec':>8}")
    print("-" * 86)
    micro_t = micro_f = micro_p = 0
    macro_f1 = []
    seg_recalls = []
    for c in COLUNAS_PII:
        rec, prec, f1, acc = met(c)
        micro_t += tp[c]; micro_f += fn[c]; micro_p += fp[c]
        macro_f1.append(f1)
        g = gold_tot[c]
        segrec = (1 - leak[c] / g) if g else 0.0
        seg_recalls.append(segrec)
        print(f"{c:<36}{g:>5}{tp[c]:>5}{fn[c]:>5}{fp[c]:>6}"
              f"{rec:>7.2f}{prec:>7.2f}{f1:>7.2f}{acc:>7.2f}{segrec:>8.2f}")

    mrec = micro_t / (micro_t + micro_f) if micro_t + micro_f else 0.0
    mprec = micro_t / (micro_t + micro_p) if micro_t + micro_p else 0.0
    mf1 = 2 * mrec * mprec / (mrec + mprec) if mrec + mprec else 0.0
    macro = sum(macro_f1) / len(macro_f1) if macro_f1 else 0.0
    seg_global = sum(seg_recalls) / len(seg_recalls) if seg_recalls else 0.0
    print("-" * 86)
    print(f"{'MICRO':<36}{'':>5}{micro_t:>5}{micro_f:>5}{micro_p:>6}"
          f"{mrec:>7.2f}{mprec:>7.2f}{mf1:>7.2f}")
    print(f"{'MACRO-F1':<36}{macro:>58.2f}")
    print(f"\n[SEGURANÇA] recall de segurança macro (não-vazamento): {seg_global:.3f}")
    print(f"[higiene] códigos institucionais (fora do PII): removidos "
          f"{inst_removidos}/{inst_total}")

    lat_med = sum(latencias)/max(1, len(latencias))
    print(f"[custo] latência média {lat_med:.2f}s/reg | "
          f"prompt_tok médio {sum(ptoks)/max(1,len(ptoks)):.0f} | "
          f"compl_tok médio {sum(ctoks)/max(1,len(ctoks)):.0f}")

    # -------------------- CSVs --------------------
    with open(os.path.join(res_dir, f'{tag}_por_tipo.csv'), 'w',
              encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['tipo', 'gold', 'tp', 'fn', 'fp', 'recall', 'precision',
                    'f1', 'accuracy', 'recall_seguranca'])
        for c in COLUNAS_PII:
            rec, prec, f1, acc = met(c)
            g = gold_tot[c]
            segrec = (1 - leak[c] / g) if g else 0.0
            w.writerow([c, g, tp[c], fn[c], fp[c], f'{rec:.4f}', f'{prec:.4f}',
                        f'{f1:.4f}', f'{acc:.4f}', f'{segrec:.4f}'])

    with open(os.path.join(res_dir, f'{tag}_matriz_erros.csv'), 'w',
              encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['tipo', 'fn_total', 'vazou_de_fato', 'removido_outro_tipo'])
        for c in COLUNAS_PII:
            w.writerow([c, fn[c], matriz[c]['vazou'], matriz[c]['mistype']])

    top_fn.sort(reverse=True)
    with open(os.path.join(res_dir, f'{tag}_top_fn.csv'), 'w',
              encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['seq', 'n_fn'])
        for nfn, seq in top_fn[:30]:
            if nfn > 0:
                w.writerow([seq, nfn])

    # tabela comparativa (append; 1 linha por config)
    comp = os.path.join(res_dir, 'tabela_comparativa.csv')
    novo = not os.path.exists(comp)
    with open(comp, 'a', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        if novo:
            w.writerow(['config', 'n', 'micro_recall', 'micro_prec', 'micro_f1',
                        'macro_f1', 'recall_seguranca_macro',
                        'recall_nome', 'recall_local', 'recall_doc',
                        'latencia_med_s', 'compl_tok_med'])
        w.writerow([tag, n, f'{mrec:.4f}', f'{mprec:.4f}', f'{mf1:.4f}',
                    f'{macro:.4f}', f'{seg_global:.4f}',
                    f'{met("Nome(s)")[0]:.4f}', f'{met("Localizações (s)")[0]:.4f}',
                    f'{met("Documento(s)")[0]:.4f}',
                    f'{lat_med:.3f}', f'{sum(ctoks)/max(1,len(ctoks)):.0f}'])

    print(f"\n  CSVs em {res_dir}/  | linha adicionada em tabela_comparativa.csv")


if __name__ == '__main__':
    main()
