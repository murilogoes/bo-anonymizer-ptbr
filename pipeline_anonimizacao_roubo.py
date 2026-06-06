"""
pipeline_anonimizacao_roubo.py — Pipeline completa de anonimização (ROUBO)
==========================================================================
Dois estágios:
  Estágio 1 — regex determinístico (estagio1_regex.py)
  Estágio 2 — camada semântica PLUGÁVEL (estagio2_motores.py)

Roda a pipeline inteira sobre a planilha auditada e grava um CSV com, por
registro: texto anonimizado + as entidades REMOVIDAS por estágio/tipo (em JSON)
+ status + latência + tokens. Esse CSV é consumido por `avaliar_pipeline.py`,
que calcula as métricas contra o gabarito (assim a avaliação não precisa
re-rodar os modelos pesados).

Recursos herdados do VDM: checkpoint/retomada por configuração, validação
pós-substituição (placeholders do Estágio 1 preservados; texto não encurta
absurdamente) e log por registro.

Uso:
  python pipeline_anonimizacao_roubo.py --motor regex_only
  python pipeline_anonimizacao_roubo.py --motor ner --ner-backend spacy
  python pipeline_anonimizacao_roubo.py --motor ner --ner-backend bertimbau
  python pipeline_anonimizacao_roubo.py --motor llm --modelo qwen3:8b --amostra 150
  python pipeline_anonimizacao_roubo.py --motor hibrido --modelo qwen3:8b --ner-backend spacy
  python pipeline_anonimizacao_roubo.py --motor encadeado --ordem ner_llm --modelo qwen3:8b
  python pipeline_anonimizacao_roubo.py --motor presidio
  python pipeline_anonimizacao_roubo.py --motor llm --modelo gemma3:12b --resume

Cada configuração grava em saidas/<tag>.csv e checkpoint em checkpoint_roubo/<tag>/.
"""
import argparse
import csv
import json
import os
import re
import sys
import time

import yaml
import openpyxl

from estagio1_regex import aplicar_estagio1
from estagio2_motores import criar_motor, aplicar_substituicoes

# Placeholders do Estágio 1 que devem sobreviver à substituição do Estágio 2
RE_PH_ESTAGIO1 = re.compile(
    r'\[(?:TELEFONE|EMAIL|REDE_SOCIAL|URL|PROCESSO|PROTOCOLO|BO|INQUERITO|'
    r'OFICIO|CODIGO|CPF|RG|CNH|IMEI|REG_PROF|DATA|VIATURA|PLACA|CEP|ENDERECO)\]'
)


def tag_config(args):
    """Identificador curto da configuração (nome de arquivo/checkpoint)."""
    t = args.motor
    if args.motor == 'ner':
        t += f"_{args.ner_backend}"
    if args.motor in ('llm', 'hibrido', 'encadeado') and args.modelo:
        t += f"_{args.modelo.replace(':', '-')}"
    if args.motor == 'encadeado':
        t += f"_{args.ordem}"
    return t


def validar_pos(texto_in, texto_out):
    """Validação pós-substituição (herdada do VDM)."""
    # 1) placeholders do Estágio 1 preservados
    if len(RE_PH_ESTAGIO1.findall(texto_out)) < len(RE_PH_ESTAGIO1.findall(texto_in)):
        return False
    # 2) texto não encolheu absurdamente
    if len(texto_in) > 50 and len(texto_out) < len(texto_in) * 0.3:
        return False
    return True


def processar_registro(texto, motor):
    """Roda a pipeline inteira num registro.

    Retorna dict com texto anonimizado, entidades removidas por estágio/tipo,
    status, latência e (se LLM) tokens.
    """
    t0 = time.time()
    if not texto or not str(texto).strip():
        return {'texto': '', 'estagio1': {}, 'estagio2': _e2vazio(),
                'status': 'vazio', 'latencia_s': 0.0,
                'prompt_tokens': 0, 'completion_tokens': 0}

    texto = str(texto)
    # Estágio 1
    txt1, caps = aplicar_estagio1(texto, retornar_capturas=True)
    # Estágio 2 (sobre o texto já anonimizado pelo regex)
    ent = motor.extrair(txt1)
    # Substituição local compartilhada
    txt2, n_subs = aplicar_substituicoes(txt1, ent, numerar=True)

    status = 'ok'
    if not validar_pos(txt1, txt2):
        # fallback: mantém só o Estágio 1 (não arrisca corromper)
        txt2, ent, status = txt1, _e2vazio(), 'fallback_estagio1'

    return {
        'texto': txt2,
        'estagio1': caps,           # {placeholder: [spans]}
        'estagio2': ent,            # {nomes/locais/estabelecimentos/vulgos: [..]}
        'status': status,
        'latencia_s': round(time.time() - t0, 3),
        'prompt_tokens': getattr(motor, 'ultimo_prompt_tokens', 0),
        'completion_tokens': getattr(motor, 'ultimo_completion_tokens', 0),
    }


def _e2vazio():
    return {'nomes': [], 'locais': [], 'estabelecimentos': [], 'vulgos': []}


def main():
    ap = argparse.ArgumentParser(description='Pipeline de anonimização — ROUBO')
    ap.add_argument('--motor', default='regex_only',
                    choices=['regex_only', 'ner', 'llm', 'hibrido', 'encadeado', 'presidio'])
    ap.add_argument('--modelo', default=None, help='modelo NER/LLM (ex.: qwen3:8b)')
    ap.add_argument('--ner-backend', default='spacy', choices=['spacy', 'bertimbau'])
    ap.add_argument('--ordem', default='ner_llm', choices=['ner_llm', 'llm_ner'])
    ap.add_argument('--amostra', type=int, default=0, help='N primeiros BOs (0=todos)')
    ap.add_argument('--resume', action='store_true', help='retoma do checkpoint')
    ap.add_argument('--config', default='config.yaml')
    args = ap.parse_args()

    cfg = yaml.safe_load(open(args.config, encoding='utf-8'))
    tag = tag_config(args)
    saida_dir = cfg['paths']['saida_dir']
    ckpt_dir = os.path.join(cfg['paths']['checkpoint_dir'], tag)
    os.makedirs(saida_dir, exist_ok=True)
    os.makedirs(ckpt_dir, exist_ok=True)
    saida_csv = os.path.join(saida_dir, f'{tag}.csv')
    ckpt_jsonl = os.path.join(ckpt_dir, 'progresso.jsonl')

    print("=" * 68)
    print(f"PIPELINE ROUBO | config = {tag}")
    print("=" * 68)

    # Carregar dataset
    wb = openpyxl.load_workbook(cfg['paths']['dataset'], read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header) if h}
    ti = idx[cfg['paths']['coluna_texto_entrada']]
    si = idx[cfg['paths']['coluna_seq']]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if args.amostra:
        rows = rows[:args.amostra]
    print(f"  registros: {len(rows)}")

    # Motor
    motor = criar_motor(cfg, motor=args.motor, modelo=args.modelo,
                        ordem=args.ordem, ner_backend=args.ner_backend)
    print(f"  motor: {motor.nome}")

    # Checkpoint
    processados = []
    if args.resume and os.path.exists(ckpt_jsonl):
        with open(ckpt_jsonl, encoding='utf-8') as f:
            processados = [json.loads(l) for l in f if l.strip()]
        print(f"  retomando: {len(processados)} já processados")
    inicio = len(processados)

    t0 = time.time()
    fckpt = open(ckpt_jsonl, 'a', encoding='utf-8')
    stats = {'ok': 0, 'fallback_estagio1': 0, 'vazio': 0}
    for i in range(inicio, len(rows)):
        row = rows[i]
        r = processar_registro(row[ti], motor)
        reg = {
            'seq': row[si],
            'texto_entrada': str(row[ti] or ''),
            'texto_anonimizado': r['texto'],
            'estagio1': r['estagio1'],
            'estagio2': r['estagio2'],
            'status': r['status'],
            'latencia_s': r['latencia_s'],
            'prompt_tokens': r['prompt_tokens'],
            'completion_tokens': r['completion_tokens'],
        }
        processados.append(reg)
        fckpt.write(json.dumps(reg, ensure_ascii=False) + '\n')
        fckpt.flush()
        stats[r['status']] = stats.get(r['status'], 0) + 1

        done = i + 1
        if done % 25 == 0 or done == len(rows):
            el = time.time() - t0
            rate = (done - inicio) / el if el > 0 else 0
            eta = (len(rows) - done) / rate if rate > 0 else 0
            print(f"  {done:>4}/{len(rows)} | {rate:.2f} reg/s | "
                  f"ETA {eta/60:.1f}min | ok={stats['ok']} "
                  f"fb={stats['fallback_estagio1']}", flush=True)
    fckpt.close()

    # Escrever CSV final
    with open(saida_csv, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['Seq', 'texto_entrada', 'texto_anonimizado',
                    'removidos_estagio1_json', 'removidos_estagio2_json',
                    'status', 'latencia_s', 'prompt_tokens', 'completion_tokens'])
        for reg in processados:
            w.writerow([reg['seq'], reg['texto_entrada'], reg['texto_anonimizado'],
                        json.dumps(reg['estagio1'], ensure_ascii=False),
                        json.dumps(reg['estagio2'], ensure_ascii=False),
                        reg['status'], reg['latencia_s'],
                        reg['prompt_tokens'], reg['completion_tokens']])

    el = time.time() - t0
    lat = [reg['latencia_s'] for reg in processados]
    print("-" * 68)
    print(f"  concluído em {el/60:.1f}min | {len(processados)} registros")
    print(f"  latência média: {sum(lat)/max(1,len(lat)):.2f}s/registro")
    print(f"  status: {stats}")
    print(f"  saída: {saida_csv}")
    print(f"  -> avalie com: python avaliar_pipeline.py --saida {saida_csv}")


if __name__ == '__main__':
    main()
