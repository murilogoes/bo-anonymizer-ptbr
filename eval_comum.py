"""
eval_comum.py — Lógica de avaliação compartilhada (idêntica a avaliar_pipeline.py),
exposta como biblioteca para reuso em analise_estatistica.py, metricas_utilidade.py
e conserto_gps_producao.py, SEM reexecutar modelos pesados.

Produz estatísticas POR REGISTRO (necessárias para bootstrap), reconciliáveis com
resultados_avaliacao/*_por_tipo.csv e tabela_comparativa.csv.
"""
import csv
import json
import re

import yaml
import openpyxl

from estagio1_regex import normalizar

COLUNAS_PII = [
    'Nome(s)', 'Localizações (s)', 'Documento(s)', 'Emplacamento (s)',
    'Telefone (s)', 'Email (s) e contas (ex: instagram)', 'N° da VTR',
    'Vulgo(s)/Apelido(s)',
]

CAT_ESTAGIO2 = {
    'nomes': 'Nome(s)',
    'locais': 'Localizações (s)',
    'estabelecimentos': 'Localizações (s)',
    'vulgos': 'Vulgo(s)/Apelido(s)',
}

RE_DOC_INSTITUCIONAL = re.compile(
    r'(?i)\b(?:IML|NOC|BOPM|BO|RE|processo|CNJ|boletim|of[íi]cio|inqu[ée]rito|protocolo)\b'
)


def split_gold(valor):
    if valor is None:
        return []
    out = []
    for parte in str(valor).split(';'):
        n = normalizar(parte)
        if n:
            out.append((parte.strip(), n))
    return out


def casa(gn, preds):
    for p in preds:
        if not p:
            continue
        if gn == p:
            return True
        if len(gn) >= 4 and len(p) >= 4 and (gn in p or p in gn):
            return True
    return False


def carregar_cfg(path='config.yaml'):
    return yaml.safe_load(open(path, encoding='utf-8'))


def carregar_gabarito(cfg):
    wb = openpyxl.load_workbook(cfg['paths']['dataset'], read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header) if h}
    gab = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq = row[idx[cfg['paths']['coluna_seq']]]
        gab[str(seq)] = {col: split_gold(row[idx[col]]) for col in COLUNAS_PII}
    return gab


def construir_de_para(cfg):
    ph2col = {ph.strip('[]'): col for ph, col in cfg['de_para'].items()}
    higiene = {h.strip('[]') for h in cfg['placeholders_higiene']}
    return ph2col, higiene


def predicoes(e1, e2, ph2col):
    """e1/e2: dicts já carregados do JSON. Retorna (pred_por_coluna, todos)."""
    pred = {c: [] for c in COLUNAS_PII}
    todos = []
    for ph, spans in e1.items():
        norm = [normalizar(s) for s in spans]
        todos += norm
        col = ph2col.get(ph)
        if col in pred:
            pred[col] += norm
    for cat, ents in e2.items():
        norm = [normalizar(s) for s in ents]
        todos += norm
        col = CAT_ESTAGIO2.get(cat)
        if col in pred:
            pred[col] += norm
    return pred, todos


def reclassificar_documento(gold_doc):
    pii, inst = [], []
    for orig, gn in gold_doc:
        (inst if RE_DOC_INSTITUCIONAL.search(orig) else pii).append((orig, gn))
    return pii, inst


def avaliar_registro(gold_reg, e1, e2, ph2col):
    """Avalia 1 BO. Retorna dict tipo -> (tp, fn, fp, leak, gold_n)."""
    pred, todos = predicoes(e1, e2, ph2col)
    out = {}
    for col in COLUNAS_PII:
        gold = gold_reg[col]
        if col == 'Documento(s)':
            gold, _inst = reclassificar_documento(gold)
        tp = fn = fp = leak = 0
        preds = pred[col]
        casados = set()
        for orig, gn in gold:
            if casa(gn, preds):
                tp += 1
                casados.update(p for p in preds if gn == p or
                               (len(gn) >= 4 and len(p) >= 4 and (gn in p or p in gn)))
            else:
                fn += 1
                if not casa(gn, todos):
                    leak += 1
        for p in set(preds):
            if p and p not in casados:
                fp += 1
        out[col] = (tp, fn, fp, leak, len(gold))
    return out


def stats_por_registro(saida_csv, gab, ph2col, e1_override=None):
    """Percorre o CSV de saída e devolve listas paralelas:
    seqs, stats (lista de dicts tipo->tupla), registros (linhas cruas).
    e1_override: dict seq -> capturas estágio 1 (para o conserto do GPS)."""
    seqs, stats, regs = [], [], []
    with open(saida_csv, encoding='utf-8') as f:
        for reg in csv.DictReader(f, delimiter=';'):
            seq = str(reg['Seq'])
            if seq not in gab:
                continue
            e1 = json.loads(reg['removidos_estagio1_json'] or '{}')
            if e1_override is not None:
                e1 = e1_override.get(seq, e1)
            e2 = json.loads(reg['removidos_estagio2_json'] or '{}')
            seqs.append(seq)
            stats.append(avaliar_registro(gab[seq], e1, e2, ph2col))
            regs.append(reg)
    return seqs, stats, regs


def agregar(stats):
    """Agrega lista de stats por registro -> métricas globais (mesmas do harness)."""
    tp = {c: 0 for c in COLUNAS_PII}
    fn = {c: 0 for c in COLUNAS_PII}
    fp = {c: 0 for c in COLUNAS_PII}
    leak = {c: 0 for c in COLUNAS_PII}
    gold = {c: 0 for c in COLUNAS_PII}
    for st in stats:
        for c, (t, f, p, l, g) in st.items():
            tp[c] += t; fn[c] += f; fp[c] += p; leak[c] += l; gold[c] += g
    mt, mf, mp = sum(tp.values()), sum(fn.values()), sum(fp.values())
    mrec = mt / (mt + mf) if mt + mf else 0.0
    mprec = mt / (mt + mp) if mt + mp else 0.0
    mf1 = 2 * mrec * mprec / (mrec + mprec) if mrec + mprec else 0.0
    segs = [(1 - leak[c] / gold[c]) if gold[c] else 0.0 for c in COLUNAS_PII]
    seg_macro = sum(segs) / len(segs)
    return {'tp': tp, 'fn': fn, 'fp': fp, 'leak': leak, 'gold': gold,
            'micro_recall': mrec, 'micro_prec': mprec, 'micro_f1': mf1,
            'recall_seguranca_macro': seg_macro}
